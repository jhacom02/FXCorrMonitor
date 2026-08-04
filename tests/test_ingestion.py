"""Tests for Excel ingestion and SQLite UPSERT behavior."""

from __future__ import annotations

import sys
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import pytest

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from config.instruments import INSTRUMENT_BY_ID, TARGET_ID
from src.database import init_db, load_market_data, upsert_market_data
from src.ingestion import (
    IngestionError,
    _normalize_sheet_name,
    _normalize_source_code,
    build_sheet_catalog,
    ingest_excel,
    read_instrument_sheet,
    resolve_sheet_for_instrument,
)
from src.utils import normalize_date, parse_numeric, utc_now_iso


def test_excel_serial_date_normalized():
    d = normalize_date(43832)
    assert d == date(2020, 1, 2)


def test_yyyymmdd_int_date_normalized():
    assert normalize_date(20160104) == date(2016, 1, 4)
    assert normalize_date(20260804) == date(2026, 8, 4)
    assert normalize_date("20260803") == date(2026, 8, 3)


def test_datetime_date_normalized():
    assert normalize_date(datetime(2024, 5, 1, 15, 30)) == date(2024, 5, 1)
    assert normalize_date(pd.Timestamp("2024-05-01")) == date(2024, 5, 1)
    assert normalize_date("2024-05-01") == date(2024, 5, 1)


def test_parse_numeric_string():
    assert parse_numeric("1,234.5") == 1234.5
    assert parse_numeric("(100)") == -100.0
    assert parse_numeric(None) is None


def test_normalize_sheet_and_code():
    assert _normalize_sheet_name("달러인덱스 Dollars_1") == "달러인덱스 Dollars"
    assert _normalize_sheet_name("KOSPI_5") == "KOSPI"
    assert _normalize_source_code(1.0) == "1"
    assert _normalize_source_code("DOLLAR") == "DOLLAR"


def test_sqlite_upsert_no_duplicate(tmp_path: Path):
    db = tmp_path / "t.db"
    init_db(db)
    from src.database import upsert_instruments

    upsert_instruments(
        [
            {
                "instrument_id": "USDKRW",
                "display_name": "USDKRW",
                "source_sheet": "s",
                "source_code": "c",
                "source_column": "현재가",
                "data_type": "price",
                "transformation": "log_return",
                "alignment": "same_day",
                "active": 1,
                "updated_at": utc_now_iso(),
            }
        ],
        db,
    )
    rows1 = [
        {
            "date": "2024-01-02",
            "instrument_id": "USDKRW",
            "raw_value": 1300.0,
            "source_file": "a.xlsx",
            "source_sheet": "s",
            "source_column": "현재가",
            "loaded_at": utc_now_iso(),
        }
    ]
    upsert_market_data(rows1, db)
    rows2 = [
        {
            "date": "2024-01-02",
            "instrument_id": "USDKRW",
            "raw_value": 1301.5,
            "source_file": "b.xlsx",
            "source_sheet": "s",
            "source_column": "현재가",
            "loaded_at": utc_now_iso(),
        }
    ]
    upsert_market_data(rows2, db)
    df = load_market_data(db, instrument_ids=["USDKRW"])
    assert len(df) == 1
    assert float(df.iloc[0]["raw_value"]) == 1301.5


def _write_minimal_excel(path: Path, include_usdkrw: bool = True, with_suffix: bool = False) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    usd_title = "서울외환(기업용) USDKRW 스팟_0" if with_suffix else "서울외환(기업용) USDKRW 스팟"
    dxy_title = "달러인덱스 Dollars_1" if with_suffix else "달러인덱스 Dollars"

    if include_usdkrw:
        ws = wb.active
        ws.title = usd_title
        ws.append(["시작", datetime(2024, 1, 1), "종료", datetime(2024, 1, 10), "종목코드", "USDSP_SMBCC_EXT"])
        ws.append(["서울외환(기업용) USDKRW 스팟"])
        ws.append(["일자", "현재가"])
        ws.append([datetime(2024, 1, 2), 1295])
        ws.append([datetime(2024, 1, 3), 1300])
        ws.append([datetime(2024, 1, 3), 1302])
    else:
        ws = wb.active
        ws.title = "Other"

    ws2 = wb.create_sheet(dxy_title)
    ws2.append(["시작", datetime(2024, 1, 1), "종료", datetime(2024, 1, 10), "종목코드", "DOLLAR"])
    ws2.append(["달러인덱스 Dollars"])
    ws2.append(["일자", "KR_MID_Close"])
    ws2.append([43832, 101])
    ws2.append([datetime(2024, 1, 2), 104])
    ws2.append([datetime(2024, 1, 3), 105])

    wb.save(path)


def test_ingest_requires_usdkrw(tmp_path: Path):
    xlsx = tmp_path / "no_usd.xlsx"
    _write_minimal_excel(xlsx, include_usdkrw=False)
    db = tmp_path / "t.db"
    with pytest.raises(IngestionError):
        ingest_excel(xlsx, db_path=db, replace=True)


def test_ingest_dedupes_dates_and_continues(tmp_path: Path):
    xlsx = tmp_path / "ok.xlsx"
    _write_minimal_excel(xlsx, include_usdkrw=True)
    db = tmp_path / "t.db"
    result = ingest_excel(xlsx, db_path=db, replace=True)
    assert result["status"] == "success"
    usd = load_market_data(db, instrument_ids=["USDKRW"])
    row = usd[usd["date"] == "2024-01-03"]
    assert len(row) == 1
    assert float(row.iloc[0]["raw_value"]) == 1302.0


def test_resolve_sheet_by_code_with_suffix(tmp_path: Path):
    xlsx = tmp_path / "sfx.xlsx"
    _write_minimal_excel(xlsx, include_usdkrw=True, with_suffix=True)
    catalog, warns = build_sheet_catalog(xlsx)
    assert not any("여러 시트" in w for w in warns)
    inst = INSTRUMENT_BY_ID["USDKRW"]
    actual, _ = resolve_sheet_for_instrument(inst, catalog)
    assert actual == "서울외환(기업용) USDKRW 스팟_0"
    cleaned, w, sheet = read_instrument_sheet(xlsx, inst, catalog=catalog)
    assert sheet == "서울외환(기업용) USDKRW 스팟_0"
    assert not cleaned.empty
    assert float(cleaned.iloc[-1]["raw_value"]) == 1302.0


def test_ambiguous_code_falls_back_to_sheet_name(tmp_path: Path):
    from openpyxl import Workbook

    xlsx = tmp_path / "amb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "달러인덱스 Dollars_1"
    ws.append(["시작", datetime(2024, 1, 1), "종료", datetime(2024, 1, 10), "종목코드", "DOLLAR"])
    ws.append(["달러인덱스 Dollars"])
    ws.append(["일자", "KR_MID_Close"])
    ws.append([datetime(2024, 1, 2), 104])
    ws2 = wb.create_sheet("기타_2")
    ws2.append(["시작", datetime(2024, 1, 1), "종료", datetime(2024, 1, 10), "종목코드", "DOLLAR"])
    ws2.append(["기타"])
    ws2.append(["일자", "KR_MID_Close"])
    ws2.append([datetime(2024, 1, 2), 999])
    wb.save(xlsx)

    catalog, warns = build_sheet_catalog(xlsx)
    assert any("DOLLAR" in w and "여러 시트" in w for w in warns)
    inst = INSTRUMENT_BY_ID["DXY"]
    actual, rwarn = resolve_sheet_for_instrument(inst, catalog)
    assert actual == "달러인덱스 Dollars_1"
    assert any("모호" in w for w in rwarn)


def test_f_net_and_ktb_instrument_ids():
    assert "F_NET" in INSTRUMENT_BY_ID
    assert "KOSPI_FOREIGN_NET" not in INSTRUMENT_BY_ID
    assert INSTRUMENT_BY_ID["KTB3Y"].source_code == "BONDKSDCAL11"
    assert INSTRUMENT_BY_ID["KTB3Y"].source_column == "대표수익률"
    assert INSTRUMENT_BY_ID["KTB10Y"].source_code == "BONDKSDCAL13"
    assert not hasattr(INSTRUMENT_BY_ID["KTB3Y"], "note")
